from math import exp
from openpyxl import load_workbook
import matplotlib
import matplotlib.pyplot as plt


matplotlib.use('TkAgg')


# def count_d(z):
#     return exp(-exp(-z))
#
#
# def count_z(x):
#     return (x - 5) / (6 - 5)
#
#
# X = []
# Y = []
# arr = np.arange(0, 10.2, 0.2)
# for i in arr:
#     z = count_z(i)
#     print(i, count_d(z))
#     X.append(i)
#     Y.append(count_d(z))
#
#
# fig = plt.figure()
#
# graph1 = plt.plot(X, Y)
# print('Plot: ', len(graph1), graph1)
#
# grid1 = plt.grid(True)
#
# plt.show()
# plt.close()


def f_harrington_unilateral(x_low, x_up, x):
    z = (x - x_low) / (x_up - x_low)
    d = 1 / (1 + exp(-3 * z * z))
    return d


def f_harrington_bilateral(x_low, x_up, x):
    z = (x - x_up) / (x_low - x_up)
    d = exp(-3 * z * z)
    return d

wb = load_workbook('projects.xlsx')
D_array = []

for i in wb.sheetnames:
    sheet = wb[i]

    values_x_low = []
    values_x_up = []
    values_x = []
    values_function_type = []

    j = 2
    while sheet[f'C{j}'].value and sheet[f'D{j}'].value and sheet[f'E{j}'].value and sheet[f'F{j}'].value:
        values_x_low.append(sheet[f'C{j}'].value)
        values_x_up.append(sheet[f'D{j}'].value)
        values_x.append(sheet[f'E{j}'].value)
        values_function_type.append(int(sheet[f'F{j}'].value))
        j += 1

    d_array = []
    for j in range(len(values_x)):
        if values_function_type[j] == 1:
            d_array.append(f_harrington_unilateral(float(values_x_low[j]), float(values_x_up[j]), float(values_x[j])))
        elif values_function_type[j] == 2:
            d_array.append(f_harrington_bilateral(float(values_x_low[j]), float(values_x_up[j]), float(values_x[j])))

    D = 1  # узагальнення
    for j in d_array:
        D = D * j
    D = D ** (1 / len(d_array))

    print('По критеріям:', d_array)
    print(f'Узагальнена оцінка для проекту {i}:', D)
    D_array.append(D)


x = wb.sheetnames
y = D_array

fig, ax = plt.subplots()

ax.bar(x, y)

ax.set_facecolor('seashell')
fig.set_facecolor('floralwhite')
fig.set_figwidth(12)
fig.set_figheight(6)

plt.show()
plt.close()
